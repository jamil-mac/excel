from django.shortcuts import render, redirect
import openpyxl
from django.views.decorators.csrf import csrf_exempt
from openpyxl.workbook import Workbook

from main.models import UserModel

import telebot

import io

from django.http import HttpResponse

bot = telebot.TeleBot('6585354812:AAE7RpqTwzYVDaP1rJHDj-ulrlOvtVS_d9c')

URL = 'https://api.telegram.org/bot6585354812:AAE7RpqTwzYVDaP1rJHDj-ulrlOvtVS_d9c'


def import_data(request):
    if request.method == 'POST':
        file = request.FILES.get('excel')
        wb = openpyxl.load_workbook(file)
        sheet = wb.active
        users = UserModel.objects.all()

        for user in users:
            cell1 = sheet.cell(row=1, column=1)
            cell2 = sheet.cell(row=2, column=2)

            cell1.value = user.full_name
            cell2.value = user.event_name

            print(
                cell1.value,
                cell2.value,
            )
        wb.save(file)
        return redirect('/import/')
    return render(request, 'excel.html')


def export_data(request):
    wb = Workbook()
    sheet = wb.active
    users = UserModel.objects.all()

    for user in users:
        sheet.append([user.full_name, user.event_name])
    name = UserModel.objects.last()
    filename = name.event_name

    chat_id = 58939309
    send_excel_file(chat_id, wb, filename)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}.xlsx'

    wb.save(response)

    return response


@csrf_exempt
def web_hook_view(request):
    """ setting webhook """
    if request.method == "POST":
        bot.process_new_updates([telebot.types.Update.de_json(request.body.decode('utf-8'))])
        return HttpResponse('ok', status=200)
    return HttpResponse('ok', status=200)


@bot.message_handler(commands=['start'])  # /start
def start(message):
    bot.reply_to(message, "Moto moto")
    print(message.chat.id)


def send_excel_file(chat_id, wb, filename):
    with io.BytesIO() as output:
        wb.save(output)
        output.seek(0)
        bot.send_document(chat_id, output, visible_file_name=filename + '.xlsx')

